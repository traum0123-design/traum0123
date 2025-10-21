from __future__ import annotations

import datetime as dt
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from .schema import DEFAULT_COLUMNS  # re-exported default columns for consumers


__all__ = ["build_salesmap_workbook", "DEFAULT_COLUMNS"]


def _normalize_label_text(label: str) -> str:
    s = (label or "").strip()
    return "".join(s.split())


def build_salesmap_workbook(
    *,
    company_slug: str,
    year: int,
    month: int,
    rows: list[dict],
    all_columns: Iterable[tuple[str, str, str]],
    group_prefs: dict[str, str] | None = None,
    alias_prefs: dict[str, str] | None = None,
) -> BytesIO:
    group_prefs = group_prefs or {}
    alias_prefs = alias_prefs or {}

    EXCLUDED_META = {
        "사원코드","사원명","부서","직급",
        "입사일","퇴사일","휴직일","휴직종료일","월 시작일","월 말일","월 총 일수","근무일수",
    }

    PREFERRED_EARN_ORDER = [
        "기본급", "월급여", "상여", "식대", "자가운전보조금", "시간외수당", "연장근로수당", "현장수당", "직급수당", "연차수당", "기타수당",
    ]
    PREFERRED_DEDUCT_ORDER = [
        "국민연금", "건강보험", "고용보험", "장기요양보험료", "소득세", "지방소득세", "각종상환공제", "학자금상환액",
        "건강보험정산", "장기요양보험정산", "고용보험정산", "고용보험 연말정산", "건강보험 연말정산", "요양보험 연말정산",
    ]

    def classify_group(label: str) -> str:
        name = str(label)
        earn_kw = ["수당", "식대", "보조", "상여", "기본급", "월급여"]
        deduct_kw = ["공제", "세", "연금", "보험", "상환", "정산"]
        if any(k in name for k in earn_kw):
            return "earn"
        if any(k in name for k in deduct_kw):
            return "deduct"
        return "earn"

    def total_by_field(rows_list: list[dict], field: str) -> int:
        s = 0
        for r in rows_list:
            v = r.get(field, 0)
            try:
                if v in (None, ""):
                    continue
                s += int(float(str(v).replace(",", "").strip()))
            except Exception:
                pass
        return s

    earn_fields: list[tuple[str, str]] = []
    deduct_fields: list[tuple[str, str]] = []
    seen_earn: set[str] = set()
    seen_deduct: set[str] = set()
    for field, label, typ in all_columns:
        if typ != "number":
            continue
        if field in EXCLUDED_META or label in EXCLUDED_META:
            continue
        disp_label = (alias_prefs.get(field) or label)
        pref_grp = (group_prefs.get(field) or "").strip()
        if pref_grp == "none":
            continue
        if pref_grp in ("earn", "deduct"):
            key = _normalize_label_text(disp_label or label)
            if pref_grp == "earn":
                if key in seen_earn:
                    continue
                seen_earn.add(key)
                earn_fields.append((field, disp_label))
            else:
                if key in seen_deduct:
                    continue
                seen_deduct.add(key)
                deduct_fields.append((field, disp_label))
            continue
        # heuristic fallback
        grp = classify_group(disp_label)
        if total_by_field(rows, field) == 0:
            continue
        key = _normalize_label_text(disp_label or label)
        if grp == "earn":
            if key in seen_earn:
                continue
            seen_earn.add(key)
            earn_fields.append((field, disp_label))
        else:
            if key in seen_deduct:
                continue
            seen_deduct.add(key)
            deduct_fields.append((field, disp_label))

    def sort_by_preference(items: list[tuple[str, str]], preferred: list[str]):
        pref_index = {name: i for i, name in enumerate(preferred)}
        return sorted(items, key=lambda x: (pref_index.get(x[1], 10_000), x[1]))

    earn_fields = sort_by_preference(earn_fields, PREFERRED_EARN_ORDER)
    deduct_fields = sort_by_preference(deduct_fields, PREFERRED_DEDUCT_ORDER)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    head_fill = PatternFill("solid", fgColor="F2F3F5")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    left_fixed = ["사원코드", "사원명", "부서", "직급"]
    earn_labels = [lbl for _, lbl in earn_fields]
    deduct_labels = [lbl for _, lbl in deduct_fields]

    row1 = left_fixed[:] + ([] if not earn_labels else ["수당"] + [""] * (len(earn_labels) - 1)) + ["지급액계"] \
           + ([] if not deduct_labels else ["공제"] + [""] * (len(deduct_labels) - 1)) + ["공제액계", "차인지급액"]
    ws.append(row1)
    row2 = left_fixed[:] + earn_labels + ["지급액계"] + deduct_labels + ["공제액계", "차인지급액"]
    ws.append(row2)

    def col_to_letter(cidx: int) -> str:
        letters = ""
        while cidx:
            cidx, rem = divmod(cidx - 1, 26)
            letters = chr(65 + rem) + letters
        return letters

    start_allow = len(left_fixed) + 1
    end_allow = start_allow + max(len(earn_labels), 0) - 1
    if end_allow >= start_allow:
        end_allow_total = end_allow + 1
        ws.merge_cells(f"{col_to_letter(start_allow)}1:{col_to_letter(end_allow_total)}1")
    start_deduct = (len(left_fixed) + max(len(earn_labels), 0) + 1) + 1
    end_deduct = start_deduct + max(len(deduct_labels), 0) - 1
    if end_deduct >= start_deduct:
        end_deduct_total = end_deduct + 1
        ws.merge_cells(f"{col_to_letter(start_deduct)}1:{col_to_letter(end_deduct_total)}1")

    max_col = len(row2)
    for r in (1, 2):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = head_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border

    import calendar

    def parse_date_flex(val):
        if not val:
            return None
        if isinstance(val, dt.date):
            return val
        s = str(val).strip()
        try:
            return dt.date.fromisoformat(s)
        except Exception:
            pass
        # fallback best-effort
        import re
        parts = [p for p in re.split(r"[^0-9]", s) if p]
        if len(parts) >= 3:
            try:
                y, m, d = map(int, parts[:3])
                return dt.date(y, m, d)
            except Exception:
                return None
        return None

    def month_range_for_row(r: dict):
        s_val = r.get("월 시작일")
        e_val = r.get("월 말일")
        s_date = parse_date_flex(s_val)
        e_date = parse_date_flex(e_val)
        if not s_date or not e_date or s_date > e_date:
            first = dt.date(year, month, 1)
            last = dt.date(year, month, calendar.monthrange(year, month)[1])
            return first, last
        return s_date, e_date

    def overlap_days(a1: dt.date, a2: dt.date, b1: dt.date, b2: dt.date) -> int:
        s = max(a1, b1)
        e = min(a2, b2)
        if e < s:
            return 0
        return (e - s).days + 1

    def proration_factor(r: dict) -> tuple[int, int]:
        ms, me = month_range_for_row(r)
        total = (me - ms).days + 1
        join = parse_date_flex(r.get("입사일"))
        leave = parse_date_flex(r.get("퇴사일"))
        act_s = max(ms, join) if join else ms
        act_e = min(me, leave) if leave else me
        if act_e < act_s:
            return 0, total
        days = (act_e - act_s).days + 1
        leave_s = parse_date_flex(r.get("휴직일"))
        leave_e = parse_date_flex(r.get("휴직종료일")) or me if leave_s else None
        if leave_s:
            d = overlap_days(act_s, act_e, max(ms, leave_s), min(me, leave_e))
            days = max(0, days - d)
        return days, total

    def get_num(row: dict, field: str) -> int:
        try:
            val = row.get(field, 0)
            if val in (None, ""):
                return 0
            return int(float(str(val).replace(",", "").strip()))
        except Exception:
            return 0

    numeric_cols_idx: list[int] = []
    for idx, label in enumerate(row2, start=1):
        if label in (earn_labels + ["지급액계"] + deduct_labels + ["공제액계", "차인지급액"]):
            numeric_cols_idx.append(idx)

    for r in rows:
        lvals = [
            r.get("사원코드", ""),
            r.get("사원명", ""),
            r.get("부서", ""),
            r.get("직급", ""),
        ]
        pay_days, tot_days = proration_factor(r)
        factor = (pay_days / tot_days) if tot_days > 0 else 0.0
        earn_vals = []
        for f, lbl in earn_fields:
            base = get_num(r, f)
            if "상여" in str(lbl):
                earn_vals.append(base)
            else:
                earn_vals.append(int(base * factor))
        earn_total = sum(earn_vals)
        deduct_vals = [get_num(r, f) for f, _ in deduct_fields]
        deduct_total = sum(deduct_vals)
        net = earn_total - deduct_total
        ws.append(lvals + earn_vals + [earn_total] + deduct_vals + [deduct_total, net])

    nfmt = "#,##0"
    for rr in range(3, ws.max_row + 1):
        for cc in numeric_cols_idx:
            cell = ws.cell(row=rr, column=cc)
            cell.number_format = nfmt
            cell.border = border
        for cc in range(1, len(left_fixed) + 1):
            ws.cell(row=rr, column=cc).border = border

    if ws.max_row >= 3:
        total_row_idx = ws.max_row + 1
        ws.append([])
        def sum_col(col_idx: int) -> int:
            s = 0
            for rr in range(3, total_row_idx):
                v = ws.cell(row=rr, column=col_idx).value or 0
                try:
                    s += int(v)
                except Exception:
                    try:
                        s += int(float(str(v).replace(",", "")))
                    except Exception:
                        pass
            return s

        values = ["합계", "", "", ""]
        for i, _ in enumerate(earn_labels, start=0):
            col_idx = len(left_fixed) + 1 + i
            values.append(sum_col(col_idx))
        allow_total = sum_col(len(left_fixed) + 1 + len(earn_labels))
        values.append(allow_total)
        allow_total_idx = len(values) - 1
        start_d = len(left_fixed) + 2 + len(earn_labels)
        for i, _ in enumerate(deduct_labels, start=0):
            col_idx = start_d + i
            values.append(sum_col(col_idx))
        deduct_total = sum_col(start_d + len(deduct_labels))
        values.append(deduct_total)
        deduct_total_idx = len(values) - 1
        net_total = (values[allow_total_idx] if allow_total_idx < len(values) else 0) - (values[deduct_total_idx] if deduct_total_idx < len(values) else 0)
        values.append(net_total)

        for i, v in enumerate(values, start=1):
            ws.cell(row=total_row_idx, column=i, value=v)
        for c in range(1, len(values) + 1):
            cell = ws.cell(row=total_row_idx, column=c)
            cell.font = bold
            cell.border = border
            if c in numeric_cols_idx:
                cell.number_format = nfmt

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            sval = str(val) if val is not None else ""
            max_len = max(max_len, len(sval))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(32, max(8, max_len + 2))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
