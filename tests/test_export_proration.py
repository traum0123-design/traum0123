from __future__ import annotations

import io
import datetime as dt

from openpyxl import load_workbook

from core.exporter import build_salesmap_workbook


def test_exporter_proration_with_join_leave_and_suspension():
    # 2024-05 has 31 days. Active 11~30 inclusive (20 days), leave 20~22 (3 days) -> paid 17 days
    year, month = 2024, 5
    rows = [
        {
            "사원코드": "E01",
            "사원명": "홍길동",
            "입사일": dt.date(2024, 5, 11),
            "퇴사일": dt.date(2024, 5, 30),
            "휴직일": dt.date(2024, 5, 20),
            "휴직종료일": dt.date(2024, 5, 22),
            "기본급": 3_100_000,  # equals 100,000 per day in May
            "상여": 310_000,  # bonuses are not prorated
            "소득세": 0,
            "지방소득세": 0,
        }
    ]
    all_columns = [
        ("사원코드", "사원코드", "text"),
        ("사원명", "사원명", "text"),
        ("입사일", "입사일", "date"),
        ("퇴사일", "퇴사일", "date"),
        ("휴직일", "휴직일", "date"),
        ("휴직종료일", "휴직종료일", "date"),
        ("기본급", "기본급", "number"),
        ("상여", "상여", "number"),
        ("소득세", "소득세", "number"),
        ("지방소득세", "지방소득세", "number"),
    ]
    group_prefs = {"기본급": "earn", "상여": "earn", "소득세": "deduct", "지방소득세": "deduct"}

    bio: io.BytesIO = build_salesmap_workbook(
        company_slug="demo",
        year=year,
        month=month,
        rows=rows,
        all_columns=all_columns,
        group_prefs=group_prefs,
        alias_prefs={},
    )
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active

    # Second header row contains labels; find column indices
    header = [c.value for c in ws[2]]
    col_basic = header.index("기본급") + 1
    col_bonus = header.index("상여") + 1

    # Data row is row 3
    prorated_basic = ws.cell(row=3, column=col_basic).value
    bonus_amount = ws.cell(row=3, column=col_bonus).value

    assert prorated_basic == 1_700_000  # 17/31 of 3,100,000
    assert bonus_amount == 310_000  # not prorated

