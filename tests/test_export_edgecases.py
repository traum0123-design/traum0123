from __future__ import annotations

import io

from openpyxl import load_workbook

from core.exporter import build_salesmap_workbook


def test_exporter_handles_decimals_negatives_and_blanks():
    rows = [
        {"사원코드": "E1", "사원명": "A", "기본급": 2000000, "소득세": 100000.5, "지방소득세": 10000},
        {"사원코드": "E2", "사원명": "B", "기본급": "1,500,000", "소득세": -5000, "지방소득세": 0},
        {"사원코드": "", "사원명": "", "기본급": "", "소득세": "", "지방소득세": ""},  # blank row ignored in totals
    ]
    all_columns = [
        ("사원코드", "사원코드", "text"),
        ("사원명", "사원명", "text"),
        ("기본급", "기본급", "number"),
        ("소득세", "소득세", "number"),
        ("지방소득세", "지방소득세", "number"),
    ]
    bio: io.BytesIO = build_salesmap_workbook(
        company_slug="demo",
        year=2024,
        month=5,
        rows=rows,
        all_columns=all_columns,
        group_prefs={"기본급": "earn", "소득세": "deduct", "지방소득세": "deduct"},
        alias_prefs={},
    )
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active
    # Totals row is last row; verify numbers are integers and totals make sense
    total_row = ws[ws.max_row]
    values = [c.value for c in total_row]
    # 지급액계(earn total) at column after earn labels
    assert isinstance(values[5], int)
    assert values[5] >= 0
    # 공제액계(deduct total)
    assert isinstance(values[-2], int)
    # 차인지급액(net) = earn - deduct
    assert isinstance(values[-1], int)
    assert values[-1] == values[5] - values[-2]

