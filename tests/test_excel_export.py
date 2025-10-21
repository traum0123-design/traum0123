from __future__ import annotations

from openpyxl import load_workbook

from core.exporter import build_salesmap_workbook


def test_build_salesmap_workbook_generates_expected_headers():
    rows = [
        {
            "사원코드": "E001",
            "사원명": "홍길동",
            "기본급": 2_000_000,
            "식대": 200_000,
            "국민연금": 90_000,
            "소득세": 100_000,
            "지방소득세": 10_000,
        },
    ]
    all_columns = [
        ("사원코드", "사원코드", "text"),
        ("사원명", "사원명", "text"),
        ("기본급", "기본급", "number"),
        ("식대", "식대", "number"),
        ("국민연금", "국민연금", "number"),
        ("소득세", "소득세", "number"),
        ("지방소득세", "지방소득세", "number"),
    ]
    group_prefs = {"기본급": "earn", "식대": "earn", "국민연금": "deduct", "소득세": "deduct", "지방소득세": "deduct"}

    bio = build_salesmap_workbook(
        company_slug="test-co",
        year=2024,
        month=5,
        rows=rows,
        all_columns=all_columns,
        group_prefs=group_prefs,
        alias_prefs={},
    )

    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active

    header_first = [cell.value for cell in ws[1]]
    header_second = [cell.value for cell in ws[2]]

    assert header_first[:6] == ["사원코드", "사원명", "부서", "직급", "수당", ""]
    assert header_first[-3:] == ["공제", "공제액계", "차인지급액"]

    assert header_second[:6] == ["사원코드", "사원명", "부서", "직급", "기본급", "식대"]
    assert header_second[-5:] == ["국민연금", "소득세", "지방소득세", "공제액계", "차인지급액"]

    earnings_total = ws.cell(row=3, column=header_second.index("지급액계") + 1)
    deductions_total = ws.cell(row=3, column=header_second.index("공제액계") + 1)

    assert earnings_total.data_type == 'n'
    assert deductions_total.data_type == 'n'
