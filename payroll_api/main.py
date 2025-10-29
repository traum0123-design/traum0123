from __future__ import annotations

import json
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Body, FastAPI, Cookie, Depends, Query, HTTPException, Request, UploadFile, File, Form
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, or_, and_
from sqlalchemy.exc import IntegrityError

from .database import get_db
from core.models import Base, Company, MonthlyPayroll, WithholdingCell, ExtraField, FieldPref, AuditEvent
import os
import secrets
import logging
from urllib.parse import quote
from dotenv import load_dotenv
from fastapi.openapi.utils import get_openapi
from fastapi.middleware.cors import CORSMiddleware
from fastapi import Header
from starlette.exceptions import HTTPException as StarletteHTTPException
from http import HTTPStatus
from werkzeug.security import generate_password_hash
from core.fields import cleanup_duplicate_extra_fields
from core.rate_limit import get_admin_rate_limiter
from core.settings import get_settings
from core.services.idempotency import maybe_idempotent_json, compute_body_hash
from core.services.audit import record_event
from core.alembic_utils import ensure_up_to_date
from core.logging_utils import maybe_enable_json_logging, set_request_id
from core.observability import init_sentry
from core.db import get_engine
from core.services import companies as company_service
from core.utils.cursor import encode_cursor, decode_cursor
from core.services.auth import (
    authenticate_admin,
    authenticate_company,
    token_roles,
    extract_token,
    issue_admin_token,
    issue_company_token,
)
from core.services.payroll import (
    compute_deductions as compute_deductions_service,
    compute_withholding_tax as compute_withholding_tax_service,
)
from core.services.policy import get_policy
import uuid
from .schemas import (
    AdminCompaniesResponse,
    AdminCompaniesPageResponse,
    AdminCompanyCreateResponse,
    AdminCompanyResetResponse,
    ClientLogPayload,
    CompanySummary,
    FieldAddRequest,
    FieldAddResponse,
    FieldCalcConfigRequest,
    FieldCalcConfigResponse,
    FieldCalcInclude,
    PayrollCalcRequest,
    PayrollCalcResponse,
    FieldExemptConfigRequest,
    FieldExemptConfigResponse,
    FieldExemptEntry,
    FieldGroupConfigRequest,
    FieldGroupConfigResponse,
    FieldProrateConfigRequest,
    FieldProrateConfigResponse,
    FieldInfo,
    FieldDeleteRequest,
    HealthResponse,
    PayrollRowsResponse,
    SimpleOkResponse,
    WithholdingImportResponse,
    WithholdingResponse,
    WithholdingYearsResponse,
    WithholdingCellEntry,
    AdminWithholdingCellsPageResponse,
    ExtraFieldEntry,
    AdminExtraFieldsPageResponse,
    AdminPayrollSummary,
    AdminCompanyPayrollsPageResponse,
    AdminAuditPageResponse,
    AuditEventEntry,
    AdminPolicyHistoryPageResponse,
    AdminPolicyHistoryEntry,
    UIPrefsGetResponse,
    UIPrefsPostRequest,
    UIPrefsPostResponse,
)


ADMIN_COOKIE_NAME = "admin_token"
PORTAL_COOKIE_NAME = "portal_token"


load_dotenv()  # .env 자동 로드(개발 편의)

@asynccontextmanager
async def lifespan(_: FastAPI):
    settings = get_settings()
    if settings.payroll_auto_apply_ddl:
        # Create tables if not exist (for POC). In production use Alembic.
        Base.metadata.create_all(bind=get_engine())
    else:
        logging.getLogger("payroll_api").info(
            "PAYROLL_AUTO_APPLY_DDL=0: skipping automatic DDL. Ensure Alembic migrations have been applied."
        )
    if settings.enforce_alembic_migrations:
        if settings.payroll_auto_apply_ddl:
            logging.getLogger("payroll_api").warning(
                "PAYROLL_ENFORCE_ALEMBIC=1 while PAYROLL_AUTO_APPLY_DDL=1; skipping migration check."
            )
        else:
            ensure_up_to_date(get_engine())
    yield


router = APIRouter()


def _format_error_payload(detail: object, code: Optional[str] = None) -> dict:
    if isinstance(detail, dict):
        message = detail.get("error") or detail.get("detail") or str(detail)
        code = code or detail.get("code")
    else:
        message = str(detail or "")
    payload = {"ok": False, "error": message or "error"}
    if code:
        payload["code"] = str(code)
    return payload


def register_exception_handlers(target) -> None:
    def _wants_problem_json(request: Request) -> bool:
        accept = (request.headers.get("accept") or "").lower()
        return "application/problem+json" in accept

    def _problem_payload(request: Request, status: int, detail: str | dict | None = None):
        rid = request.headers.get("x-request-id") or ""
        try:
            title = HTTPStatus(status).phrase
        except Exception:
            title = "Error"
        if isinstance(detail, dict):
            det = detail.get("detail") or detail.get("error") or detail
        else:
            det = detail or ""
        return {
            "type": "about:blank",
            "title": title,
            "status": status,
            "detail": det,
            "instance": str(request.url.path),
            "request_id": rid,
        }

    async def http_exception_handler(request: Request, exc: StarletteHTTPException):
        if _wants_problem_json(request):
            content = _problem_payload(request, exc.status_code, exc.detail)
            return JSONResponse(status_code=exc.status_code, content=content, media_type="application/problem+json")
        payload = _format_error_payload(exc.detail)
        payload["request_id"] = request.headers.get("x-request-id") or ""
        return JSONResponse(status_code=exc.status_code, content=payload)

    async def validation_exception_handler(request: Request, exc: RequestValidationError):
        if _wants_problem_json(request):
            content = _problem_payload(request, 422, {"detail": exc.errors()})
            return JSONResponse(status_code=422, content=content, media_type="application/problem+json")
        payload = {
            "ok": False,
            "error": "validation_error",
            "code": "validation_error",
            "details": exc.errors(),
            "request_id": request.headers.get("x-request-id") or "",
        }
        return JSONResponse(status_code=422, content=payload)

    target.add_exception_handler(StarletteHTTPException, http_exception_handler)
    target.add_exception_handler(RequestValidationError, validation_exception_handler)
    
    async def sa_integrity_error_handler(request: Request, exc: IntegrityError):
        if _wants_problem_json(request):
            content = _problem_payload(request, 400, {"detail": "constraint_violation"})
            return JSONResponse(status_code=400, content=content, media_type="application/problem+json")
        payload = {
            "ok": False,
            "error": "constraint_violation",
            "code": "constraint_violation",
            "request_id": request.headers.get("x-request-id") or "",
        }
        return JSONResponse(status_code=400, content=payload)

    target.add_exception_handler(IntegrityError, sa_integrity_error_handler)


# CORS for dev/proxy scenarios
@router.get("/healthz", response_model=HealthResponse)
def healthz(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"ok": True, "status": "healthy"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get('/livez', response_model=SimpleOkResponse)
def livez():
    return SimpleOkResponse()

@router.get('/readyz', response_model=HealthResponse)
def readyz(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def get_company_by_slug(db: Session, slug: str) -> Optional[Company]:
    return db.query(Company).filter(Company.slug == slug).first()


@router.get("/portal/{slug}/api/withholding", response_model=WithholdingResponse)
@router.get("/api/portal/{slug}/withholding", response_model=WithholdingResponse)
def api_withholding(
    slug: str,
    year: int = Query(..., description="연도"),
    dep: int = Query(..., description="부양가족수"),
    wage: int = Query(..., description="월보수(과세표준)"),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    tax = compute_withholding_tax_service(db, year, dep, wage)
    return {
        "ok": True,
        "year": year,
        "dep": dep,
        "wage": wage,
        "tax": int(tax),
        "local_tax": int(round((tax or 0) * 0.1)),
    }


# ------------------------------
# Client log collector
# ------------------------------

logger = logging.getLogger("payroll_api.client_log")
audit_logger = logging.getLogger("payroll_api.audit")


@router.post('/client-log', response_model=SimpleOkResponse)
async def client_log(
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
    payload: Optional[ClientLogPayload] = Body(default=None),
):
    # Accept either admin token or company token (with revocation check)
    admin_tok = extract_token(authorization, x_admin_token, None, admin_cookie)
    who: Optional[str] = None
    if admin_tok and authenticate_admin(admin_tok):
        who = 'admin'
    else:
        company_tok = extract_token(authorization, x_api_token, token, portal_cookie)
        if company_tok:
            company = authenticate_company(db, None, company_tok)
            if company:
                who = f"company:{company.slug}"
    if not who:
        raise HTTPException(status_code=403, detail="forbidden")
    # Read payload
    data = payload.dict() if payload else {}
    def _clip(v, n=2000):
        try:
            s = str(v or '')
            return s if len(s) <= n else s[:n]
        except Exception:
            return ''
    level = (data.get('level') or 'error').lower()
    if level not in {'debug', 'info', 'warning', 'error', 'critical'}:
        level = 'error'
    max_stack = int(os.environ.get("CLIENT_LOG_STACK_MAX", "4000") or 4000)
    max_msg = int(os.environ.get("CLIENT_LOG_MESSAGE_MAX", "2000") or 2000)
    max_url = int(os.environ.get("CLIENT_LOG_URL_MAX", "512") or 512)
    max_ua = int(os.environ.get("CLIENT_LOG_UA_MAX", "512") or 512)
    out = {
        "ts": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "who": who,
        "lvl": level,
        "msg": _clip(data.get('message'), max_msg),
        "url": _clip(data.get('url'), max_url),
        "ua": _clip(data.get('ua'), max_ua),
        "line": data.get('line') or '',
        "col": data.get('col') or '',
        "stack": _clip(data.get('stack'), max_stack),
        "kind": data.get('kind') or 'onerror',
    }
    try:
        ip = getattr(request.client, 'host', None)
        if not ip:
            forwarded = request.headers.get('x-forwarded-for', '')
            ip = forwarded.split(',')[0].strip() if forwarded else 'unknown'
    except Exception:
        ip = 'unknown'
    max_attempts = int(os.environ.get("CLIENT_LOG_RL_MAX", "120") or 120)
    window_sec = int(os.environ.get("CLIENT_LOG_RL_WINDOW", "60") or 60)
    limiter = get_admin_rate_limiter()
    key = f"clientlog:{who}:{ip}"
    try:
        if limiter.too_many_attempts(key, window_sec, max_attempts):
            raise HTTPException(status_code=429, detail="too_many_client_logs")
    except HTTPException:
        raise
    except Exception:
        pass
    try:
        logger.info("client_log", extra={"client_log": out})
    except Exception:
        try:
            logger.info("client_log %s", json.dumps(out, ensure_ascii=False))
        except Exception:
            pass
    return SimpleOkResponse()


# ------------------------------
# Admin: Withholding table
# ------------------------------

@router.get("/admin/tax/withholding/sample", response_model=WithholdingResponse)
def admin_withholding_sample(
    year: int = Query(...),
    dep: int = Query(...),
    wage: int = Query(...),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    tax = compute_withholding_tax_service(db, year, dep, wage)
    return {"ok": True, "year": year, "dep": dep, "wage": wage, "tax": int(tax), "local_tax": int(round((tax or 0) * 0.1))}


@router.post("/admin/tax/withholding/import", response_model=WithholdingImportResponse)
async def admin_withholding_import(
    year: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    from openpyxl import load_workbook
    try:
        # Read content once and compute idempotency body hash
        content = await file.read()
        import hashlib
        body_hash = compute_body_hash({"year": int(year), "sha256": hashlib.sha256(content).hexdigest()})

        def _produce():
            import io
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            header_row_idx = None
            dep_cols = {}
            for r in range(1, 15):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                tmp = {}
                for c, v in enumerate(row_vals[1:], start=2):
                    try:
                        iv = int(str(v).strip().replace(',', ''))
                        tmp[c] = iv
                    except Exception:
                        pass
                if len(tmp) >= 2:
                    header_row_idx = r
                    dep_cols = tmp
                    break
            if not header_row_idx:
                raise ValueError("의존가족수 헤더를 찾을 수 없습니다.")
            data: list[dict[str, int]] = []
            for r in range(header_row_idx+1, ws.max_row+1):
                v = ws.cell(row=r, column=1).value
                if v is None:
                    continue
                try:
                    wage_v = int(float(str(v).replace(',', '').strip()))
                except Exception:
                    if data:
                        break
                    else:
                        continue
                for c, dep_v in dep_cols.items():
                    tv = ws.cell(row=r, column=c).value
                    try:
                        tax = int(float(str(tv).replace(',', '').strip())) if tv not in (None, "") else 0
                    except Exception:
                        tax = 0
                    data.append({"year": year, "dependents": dep_v, "wage": wage_v, "tax": tax})
            if not data:
                raise ValueError("유효한 데이터가 없습니다.")
            inserted = 0
            with db.begin():
                db.query(WithholdingCell).filter(WithholdingCell.year == year).delete(synchronize_session=False)
                db.bulk_insert_mappings(WithholdingCell, data)  # type: ignore[arg-type]
                inserted = len(data)
            return {"ok": True, "year": year, "count": inserted}, 200

        content_json, status = maybe_idempotent_json(
            db,
            request,  # type: ignore[name-defined]
            company_id=None,
            body_hash=body_hash,
            produce=_produce,
        )
        return JSONResponse(status_code=status, content=content_json)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/admin/api/withholding/years", response_model=WithholdingYearsResponse)
def admin_withholding_years(
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    try:
        rows = db.execute(text("SELECT year, COUNT(1) FROM withholding_cells GROUP BY year ORDER BY year DESC")).all()
        return {"ok": True, "years": [(int(y), int(c)) for (y, c) in rows]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ------------------------------
# Admin: Company management
# ------------------------------

@router.post("/admin/company/new", response_model=AdminCompanyCreateResponse)
async def admin_company_new(
    name: str = Form(...),
    slug: str = Form(...),
    request: Request = None,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    name = (name or '').strip()
    slug = (slug or '').strip().lower()
    if not name or not slug:
        raise HTTPException(status_code=400, detail="name/slug required")
    if db.query(Company).filter(Company.slug == slug).first():
        raise HTTPException(status_code=400, detail="slug in use")
    body_hash = compute_body_hash({"name": name, "slug": slug})

    def _produce():
        access_code = secrets.token_hex(4)
        access_hash = generate_password_hash(access_code)
        comp = Company(name=name, slug=slug, access_hash=access_hash)
        db.add(comp)
        db.commit()
        payload = CompanySummary(
            id=comp.id,
            name=comp.name,
            slug=comp.slug,
            created_at=comp.created_at.isoformat() if comp.created_at else None,
        )
        # Audit (DB)
        try:
            ip = getattr(request.client, 'host', None) or request.headers.get('x-forwarded-for', '').split(',')[0].strip() or ''
            ua = request.headers.get('user-agent', '')
            record_event(
                db,
                actor='admin',
                action='company_created',
                resource='/admin/company/new',
                company_id=comp.id,
                ip=ip,
                ua=ua,
                meta={"slug": slug, "name": name},
            )
        except Exception:
            pass
        return {"ok": True, "company": payload, "access_code": access_code}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=None,
        body_hash=body_hash,
        produce=_produce,
    )
    try:
        audit_logger.info(
            "company_created",
            extra={
                "event": "company_created",
                "slug": slug,
                "name": name,
                "status": status,
            },
        )
    except Exception:
        pass
    return JSONResponse(content=content, status_code=status)


@router.post("/admin/company/{company_id}/reset-code", response_model=AdminCompanyResetResponse)
def admin_company_reset_code(
    company_id: int,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    comp = db.get(Company, company_id)
    if not comp:
        raise HTTPException(status_code=404, detail="not found")
    body_hash = compute_body_hash({"company_id": company_id, "action": "reset_code"})

    def _produce():
        access_code = secrets.token_hex(4)
        comp.access_hash = generate_password_hash(access_code)
        db.commit()
        # Audit (DB)
        try:
            ip = getattr(request.client, 'host', None) or request.headers.get('x-forwarded-for', '').split(',')[0].strip() or ''
            ua = request.headers.get('user-agent', '')
            record_event(
                db,
                actor='admin',
                action='company_access_code_rotated',
                resource=f"/admin/company/{company_id}/reset-code",
                company_id=comp.id,
                ip=ip,
                ua=ua,
            )
        except Exception:
            pass
        return {"ok": True, "company_id": comp.id, "access_code": access_code}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=comp.id,
        body_hash=body_hash,
        produce=_produce,
    )
    try:
        audit_logger.info(
            "company_access_code_rotated",
            extra={
                "event": "company_access_code_rotated",
                "company_id": comp.id,
                "slug": comp.slug,
                "status": status,
            },
        )
    except Exception:
        pass
    return JSONResponse(content=content, status_code=status)


@router.get("/admin/companies", response_model=AdminCompaniesResponse)
def admin_companies(
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    rows = db.query(Company).order_by(Company.created_at.desc()).all()
    companies = [
        CompanySummary(
            id=c.id,
            name=c.name,
            slug=c.slug,
            created_at=c.created_at.isoformat() if c.created_at else None,
        )
        for c in rows
    ]
    return {"ok": True, "companies": companies}


@router.get("/admin/companies/page", response_model=AdminCompaniesPageResponse)
def admin_companies_page(
    limit: int = Query(50, ge=1, le=200),
    cursor: str | None = Query(default=None),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    q = db.query(Company)
    desc = (order or "desc").lower() != "asc"
    if desc:
        q = q.order_by(Company.created_at.desc(), Company.id.desc())
    else:
        q = q.order_by(Company.created_at.asc(), Company.id.asc())

    if cursor:
        try:
            cur = decode_cursor(cursor)
            cur_ts = cur.get("created_at")
            cur_id = int(cur.get("id"))
        except Exception:
            raise HTTPException(status_code=400, detail="invalid cursor")
        from datetime import datetime

        def _parse_ts(s: str) -> datetime:
            try:
                # Accept ISO 8601 with 'Z'
                if s.endswith("Z"):
                    from datetime import timezone
                    return datetime.fromisoformat(s.replace("Z", "+00:00"))
                return datetime.fromisoformat(s)
            except Exception:
                raise HTTPException(status_code=400, detail="invalid cursor timestamp")

        ts = _parse_ts(str(cur_ts))
        if desc:
            q = q.filter(or_(Company.created_at < ts, and_(Company.created_at == ts, Company.id < cur_id)))
        else:
            q = q.filter(or_(Company.created_at > ts, and_(Company.created_at == ts, Company.id > cur_id)))

    rows = q.limit(limit + 1).all()
    has_more = len(rows) > limit
    items_rows = rows[:limit]
    items = [
        CompanySummary(
            id=c.id,
            name=c.name,
            slug=c.slug,
            created_at=c.created_at.isoformat() if c.created_at else None,
        )
        for c in items_rows
    ]
    next_cur: str | None = None
    if has_more and items_rows:
        last = items_rows[-1]
        next_cur = encode_cursor({
            "id": last.id,
            "created_at": last.created_at,
            "order": order,
        })
    return AdminCompaniesPageResponse(items=items, next_cursor=next_cur, has_more=has_more)


@router.get("/admin/tax/withholding/cells", response_model=AdminWithholdingCellsPageResponse)
def admin_withholding_cells_page(
    year: int = Query(...),
    dep: int | None = Query(default=None, description="optional dependents filter"),
    limit: int = Query(100, ge=1, le=500),
    cursor: str | None = Query(default=None),
    order: str = Query("asc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    q = db.query(WithholdingCell).filter(WithholdingCell.year == int(year))
    if dep is not None:
        q = q.filter(WithholdingCell.dependents == int(dep))
    desc = (order or "asc").lower() == "desc"
    if desc:
        q = q.order_by(WithholdingCell.dependents.desc(), WithholdingCell.wage.desc(), WithholdingCell.id.desc())
    else:
        q = q.order_by(WithholdingCell.dependents.asc(), WithholdingCell.wage.asc(), WithholdingCell.id.asc())

    if cursor:
        try:
            cur = decode_cursor(cursor)
            cdep = int(cur.get("dep"))
            cwage = int(cur.get("wage"))
            cid = int(cur.get("id"))
        except Exception:
            raise HTTPException(status_code=400, detail="invalid cursor")

        if desc:
            # dependents DESC, wage DESC, id DESC
            q = q.filter(
                or_(
                    WithholdingCell.dependents < cdep,
                    and_(WithholdingCell.dependents == cdep, WithholdingCell.wage < cwage),
                    and_(WithholdingCell.dependents == cdep, WithholdingCell.wage == cwage, WithholdingCell.id < cid),
                )
            )
        else:
            q = q.filter(
                or_(
                    WithholdingCell.dependents > cdep,
                    and_(WithholdingCell.dependents == cdep, WithholdingCell.wage > cwage),
                    and_(WithholdingCell.dependents == cdep, WithholdingCell.wage == cwage, WithholdingCell.id > cid),
                )
            )

    rows = q.limit(limit + 1).all()
    has_more = len(rows) > limit
    items_rows = rows[:limit]
    items = [WithholdingCellEntry(dependents=r.dependents, wage=r.wage, tax=r.tax) for r in items_rows]
    next_cur: str | None = None
    if has_more and items_rows:
        last = items_rows[-1]
        next_cur = encode_cursor({"id": last.id, "dep": last.dependents, "wage": last.wage, "order": order, "year": year})
    return AdminWithholdingCellsPageResponse(items=items, next_cursor=next_cur, has_more=has_more)


@router.get("/admin/company/{company_id}/extra-fields/page", response_model=AdminExtraFieldsPageResponse)
def admin_extra_fields_page(
    company_id: int,
    limit: int = Query(50, ge=1, le=200),
    cursor: str | None = Query(default=None),
    order: str = Query("asc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    comp = db.get(Company, company_id)
    if not comp:
        raise HTTPException(status_code=404, detail="not found")
    q = db.query(ExtraField).filter(ExtraField.company_id == comp.id)
    desc = (order or "asc").lower() == "desc"
    if desc:
        q = q.order_by(ExtraField.position.desc(), ExtraField.id.desc())
    else:
        q = q.order_by(ExtraField.position.asc(), ExtraField.id.asc())

    if cursor:
        try:
            cur = decode_cursor(cursor)
            cpos = int(cur.get("position"))
            cid = int(cur.get("id"))
        except Exception:
            raise HTTPException(status_code=400, detail="invalid cursor")
        if desc:
            q = q.filter(or_(ExtraField.position < cpos, and_(ExtraField.position == cpos, ExtraField.id < cid)))
        else:
            q = q.filter(or_(ExtraField.position > cpos, and_(ExtraField.position == cpos, ExtraField.id > cid)))

    rows = q.limit(limit + 1).all()
    has_more = len(rows) > limit
    items_rows = rows[:limit]
    items = [
        ExtraFieldEntry(id=r.id, name=r.name, label=r.label, typ=r.typ, position=int(r.position or 0)) for r in items_rows
    ]
    next_cur: str | None = None
    if has_more and items_rows:
        last = items_rows[-1]
        next_cur = encode_cursor({"id": last.id, "position": last.position, "order": order, "company_id": company_id})
    return AdminExtraFieldsPageResponse(items=items, next_cursor=next_cur, has_more=has_more)


@router.get("/admin/company/{company_id}/payrolls/page", response_model=AdminCompanyPayrollsPageResponse)
def admin_company_payrolls_page(
    company_id: int,
    limit: int = Query(50, ge=1, le=200),
    cursor: str | None = Query(default=None),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    year: int | None = Query(default=None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    comp = db.get(Company, company_id)
    if not comp:
        raise HTTPException(status_code=404, detail="not found")
    q = db.query(MonthlyPayroll).filter(MonthlyPayroll.company_id == comp.id)
    if year is not None:
        q = q.filter(MonthlyPayroll.year == int(year))
    desc = (order or "desc").lower() != "asc"
    if desc:
        q = q.order_by(MonthlyPayroll.year.desc(), MonthlyPayroll.month.desc(), MonthlyPayroll.id.desc())
    else:
        q = q.order_by(MonthlyPayroll.year.asc(), MonthlyPayroll.month.asc(), MonthlyPayroll.id.asc())

    if cursor:
        try:
            cur = decode_cursor(cursor)
            cy = int(cur.get("year"))
            cm = int(cur.get("month"))
            cid = int(cur.get("id"))
        except Exception:
            raise HTTPException(status_code=400, detail="invalid cursor")
        if desc:
            q = q.filter(
                or_(
                    MonthlyPayroll.year < cy,
                    and_(MonthlyPayroll.year == cy, MonthlyPayroll.month < cm),
                    and_(MonthlyPayroll.year == cy, MonthlyPayroll.month == cm, MonthlyPayroll.id < cid),
                )
            )
        else:
            q = q.filter(
                or_(
                    MonthlyPayroll.year > cy,
                    and_(MonthlyPayroll.year == cy, MonthlyPayroll.month > cm),
                    and_(MonthlyPayroll.year == cy, MonthlyPayroll.month == cm, MonthlyPayroll.id > cid),
                )
            )

    rows = q.limit(limit + 1).all()
    has_more = len(rows) > limit
    items_rows = rows[:limit]
    items = [
        AdminPayrollSummary(
            id=r.id,
            year=int(r.year),
            month=int(r.month),
            is_closed=bool(getattr(r, "is_closed", False)),
            updated_at=r.updated_at.isoformat() if getattr(r, "updated_at", None) else None,
        )
        for r in items_rows
    ]
    next_cur: str | None = None
    if has_more and items_rows:
        last = items_rows[-1]
        next_cur = encode_cursor({
            "year": last.year,
            "month": last.month,
            "id": last.id,
            "order": order,
            "company_id": company_id,
        })
    return AdminCompanyPayrollsPageResponse(items=items, next_cursor=next_cur, has_more=has_more)


@router.get("/admin/company/{company_id}/impersonate-token")
def admin_impersonate_token(
    company_id: int,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    # Returns a portal token for the given company
    require_admin(authorization, x_admin_token, None, admin_cookie)
    comp = db.get(Company, company_id)
    if not comp:
        raise HTTPException(status_code=404, detail="not found")
    # Impersonation issues a company_admin role token (not app admin)
    tok = issue_company_token(db, comp, is_admin=False, roles=["company_admin"])
    try:
        audit_logger.info(
            "impersonate_token_issued",
            extra={"event": "impersonate_token_issued", "company_id": comp.id, "slug": comp.slug},
        )
    except Exception:
        pass
    # Audit (DB)
    try:
        record_event(db, actor='admin', action='impersonate_token_issued', resource=f"/admin/company/{company_id}/impersonate-token", company_id=comp.id)
    except Exception:
        pass
    return {"ok": True, "slug": comp.slug, "token": tok}


@router.get("/admin/audit", response_model=AdminAuditPageResponse)
def admin_audit_list(
    limit: int = Query(50, ge=1, le=200),
    cursor: str | None = Query(default=None),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    company_id: int | None = Query(default=None),
    actor: str | None = Query(default=None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    # Optional roles check: require 'admin' if roles present
    tok = extract_token(authorization, x_admin_token, None, admin_cookie)
    try:
        if tok and "admin" not in set(token_roles(tok, is_admin=True)):
            raise HTTPException(status_code=403, detail="forbidden")
    except Exception:
        pass
    q = db.query(AuditEvent)
    if company_id is not None:
        q = q.filter(AuditEvent.company_id == int(company_id))
    if actor:
        q = q.filter(AuditEvent.actor == str(actor))
    desc = (order or "desc").lower() != "asc"
    if desc:
        q = q.order_by(AuditEvent.ts.desc(), AuditEvent.id.desc())
    else:
        q = q.order_by(AuditEvent.ts.asc(), AuditEvent.id.asc())

    if cursor:
        try:
            cur = decode_cursor(cursor)
            cid = int(cur.get("id"))
        except Exception:
            raise HTTPException(status_code=400, detail="invalid cursor")
        if desc:
            q = q.filter(AuditEvent.id < cid)
        else:
            q = q.filter(AuditEvent.id > cid)

    rows = q.limit(limit + 1).all()
    has_more = len(rows) > limit
    items_rows = rows[:limit]
    items: list[AuditEventEntry] = []
    for r in items_rows:
        try:
            meta = json.loads(getattr(r, "meta_json", "") or "{}")
            if not isinstance(meta, dict):
                meta = {}
        except Exception:
            meta = {}
        items.append(
            AuditEventEntry(
                id=r.id,
                ts=(r.ts.isoformat().replace("+00:00", "Z") if getattr(r, "ts", None) else ""),
                actor=r.actor,
                company_id=r.company_id,
                action=r.action,
                resource=r.resource or "",
                ip=r.ip or "",
                ua=r.ua or "",
                result=r.result or "",
                meta=meta,
            )
        )
    next_cur: str | None = None
    if has_more and items_rows:
        last = items_rows[-1]
        next_cur = encode_cursor({"id": last.id, "order": order, "company_id": company_id, "actor": actor})
    return AdminAuditPageResponse(items=items, next_cursor=next_cur, has_more=has_more)


@router.get("/admin/policy")
def admin_get_policy(
    year: int = Query(...),
    company_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    pol = get_policy(db, company_id, year)
    return {"ok": True, "policy": pol}


@router.post("/admin/policy", response_model=SimpleOkResponse)
def admin_set_policy(
    request: Request,
    year: int = Query(...),
    company_id: int | None = Query(default=None),
    body: dict | None = Body(default=None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    if body is None or not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="invalid payload")

    body_hash = compute_body_hash({"company_id": company_id, "year": year, "policy": body})

    def _produce():
        from core.models import PolicySetting, PolicySettingHistory
        # Read previous policy for history
        rec = (
            db.query(PolicySetting)
            .filter(PolicySetting.company_id.is_(None) if company_id is None else PolicySetting.company_id == int(company_id))
            .filter(PolicySetting.year == int(year))
            .first()
        )
        old = {}
        if rec:
            try:
                old = json.loads(rec.policy_json or "{}")
            except Exception:
                old = {}
            rec.policy_json = json.dumps(body, ensure_ascii=False)
        else:
            rec = PolicySetting(company_id=company_id, year=int(year), policy_json=json.dumps(body, ensure_ascii=False))
            db.add(rec)
        # Write history
        try:
            hist = PolicySettingHistory(
                actor="admin",
                company_id=company_id,
                year=int(year),
                old_json=json.dumps(old, ensure_ascii=False),
                new_json=json.dumps(body, ensure_ascii=False),
            )
            db.add(hist)
        except Exception:
            pass
        db.commit()
        try:
            audit_logger.info("policy_updated", extra={"event": "policy_updated", "company_id": company_id, "year": year})
        except Exception:
            pass
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=company_id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(content=content, status_code=status)


@router.get("/admin/policy/history", response_model=AdminPolicyHistoryPageResponse)
def admin_policy_history(
    limit: int = Query(50, ge=1, le=200),
    cursor: str | None = Query(default=None),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    company_id: int | None = Query(default=None),
    year: int | None = Query(default=None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    require_admin(authorization, x_admin_token, None, admin_cookie)
    from core.models import PolicySettingHistory
    q = db.query(PolicySettingHistory)
    if company_id is not None:
        q = q.filter(PolicySettingHistory.company_id == int(company_id))
    if year is not None:
        q = q.filter(PolicySettingHistory.year == int(year))
    desc = (order or "desc").lower() != "asc"
    if desc:
        q = q.order_by(PolicySettingHistory.ts.desc(), PolicySettingHistory.id.desc())
    else:
        q = q.order_by(PolicySettingHistory.ts.asc(), PolicySettingHistory.id.asc())
    if cursor:
        try:
            cur = decode_cursor(cursor)
            cid = int(cur.get("id"))
        except Exception:
            raise HTTPException(status_code=400, detail="invalid cursor")
        if desc:
            q = q.filter(PolicySettingHistory.id < cid)
        else:
            q = q.filter(PolicySettingHistory.id > cid)
    rows = q.limit(limit + 1).all()
    has_more = len(rows) > limit
    items_rows = rows[:limit]
    items: list[AdminPolicyHistoryEntry] = []
    for r in items_rows:
        try:
            old = json.loads(getattr(r, "old_json", "") or "{}")
        except Exception:
            old = {}
        try:
            new = json.loads(getattr(r, "new_json", "") or "{}")
        except Exception:
            new = {}
        items.append(
            AdminPolicyHistoryEntry(
                id=r.id,
                ts=r.ts.isoformat().replace("+00:00", "Z") if getattr(r, "ts", None) else "",
                actor=r.actor,
                company_id=r.company_id,
                year=r.year,
                old=old,
                new=new,
            )
        )
    next_cur = None
    if has_more and items_rows:
        last = items_rows[-1]
        next_cur = encode_cursor({"id": last.id, "order": order, "company_id": company_id, "year": year})
    return AdminPolicyHistoryPageResponse(items=items, next_cursor=next_cur, has_more=has_more)


# ------------------------------
# UI preferences (company-level)
# ------------------------------

@router.get("/portal/{slug}/ui-prefs", response_model=UIPrefsGetResponse)
@router.get("/api/portal/{slug}/ui-prefs", response_model=UIPrefsGetResponse)
def api_get_ui_prefs(
    slug: str,
    keys: str | None = Query(default=None, description="Comma-separated keys"),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    wanted = None
    if keys:
        wanted = {k.strip() for k in keys.split(',') if k.strip()}
    from core.models import UISetting
    q = db.query(UISetting).filter(UISetting.company_id == company.id)
    rows = q.all()
    out: dict[str, object] = {}
    import json as _json
    for r in rows:
        if wanted and r.key not in wanted:
            continue
        try:
            out[r.key] = _json.loads(r.value_json or "{}")
        except Exception:
            out[r.key] = {}
    return UIPrefsGetResponse(values=out)


@router.post("/portal/{slug}/ui-prefs", response_model=UIPrefsPostResponse)
@router.post("/api/portal/{slug}/ui-prefs", response_model=UIPrefsPostResponse)
def api_set_ui_prefs(
    slug: str,
    payload: UIPrefsPostRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    # Allow any authenticated portal role to update UI prefs at company scope
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    vals = payload.values or {}
    import json as _json
    from core.models import UISetting
    body_hash = compute_body_hash({"values": vals})

    def _produce():
        for k, v in (vals or {}).items():
            if not k or not isinstance(k, str):
                continue
            js = _json.dumps(v or {}, ensure_ascii=False)
            rec = db.query(UISetting).filter(UISetting.company_id == company.id, UISetting.key == k).first()
            if not rec:
                rec = UISetting(company_id=company.id, key=k, value_json=js)
                db.add(rec)
            else:
                rec.value_json = js
        db.commit()
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(db, request, company_id=company.id, body_hash=body_hash, produce=_produce)
    return JSONResponse(status_code=status, content=content)


@router.post("/admin/tokens/revoke", response_model=SimpleOkResponse)
def admin_revoke_self_token(
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    tok = extract_token(authorization, x_admin_token, None, admin_cookie)
    if not tok or not authenticate_admin(tok):
        raise HTTPException(status_code=403, detail="forbidden")
    try:
        payload = None
        from core.auth import verify_admin_token
        from core.settings import get_settings
        payload = verify_admin_token(get_settings().secret_key, tok)
        if not payload:
            raise HTTPException(status_code=400, detail="invalid token")
        jti = str(payload.get("jti") or "")
        if not jti:
            raise HTTPException(status_code=400, detail="missing jti")
        from core.models import RevokedToken
        rec = RevokedToken(typ="admin", jti=jti)
        db.add(rec)
        db.commit()
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="cannot revoke token")
    return SimpleOkResponse()


@router.post("/admin/company/{company_id}/rotate-token-key", response_model=SimpleOkResponse)
def admin_rotate_company_token_key(
    company_id: int,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
):
    """Rotate company token key to revoke existing tokens (admin-only)."""
    require_admin(authorization, x_admin_token, None, admin_cookie)
    comp = db.get(Company, company_id)
    if not comp:
        raise HTTPException(status_code=404, detail="not found")

    body_hash = compute_body_hash({"company_id": company_id, "action": "rotate_token_key"})

    def _produce():
        # rotation immediately invalidates all existing company tokens
        company_service.rotate_company_token_key(db, comp)
        # Optionally also rotate access code? no, keep separate
        try:
            audit_logger.info(
                "company_token_key_rotated",
                extra={
                    "event": "company_token_key_rotated",
                    "company_id": comp.id,
                    "slug": comp.slug,
                },
            )
        except Exception:
            pass
        # Audit (DB)
        try:
            ip = getattr(request.client, 'host', None) or request.headers.get('x-forwarded-for', '').split(',')[0].strip() or ''
            ua = request.headers.get('user-agent', '')
            record_event(
                db,
                actor='admin',
                action='company_token_key_rotated',
                resource=f"/admin/company/{company_id}/rotate-token-key",
                company_id=comp.id,
                ip=ip,
                ua=ua,
            )
        except Exception:
            pass
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=comp.id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(content=content, status_code=status)


# ------------------------------
# Admin: Login (FastAPI-only usage)
# ------------------------------


@router.post("/admin/login")
def admin_login_api(request: Request, password: str = Form(...)):
    if not company_service.verify_admin_password(password):
        # Rate limit by client IP
        try:
            ip = getattr(request.client, 'host', None) or request.headers.get('x-forwarded-for', '').split(',')[0].strip() or 'unknown'
        except Exception:
            ip = 'unknown'
        try:
            max_attempts = int(os.environ.get("ADMIN_LOGIN_RL_MAX", "10") or 10)
            window_sec = int(os.environ.get("ADMIN_LOGIN_RL_WINDOW", "600") or 600)
        except Exception:
            max_attempts = 10
            window_sec = 600
        limiter = get_admin_rate_limiter()
        key = f"fastapi:{ip}"
        try:
            exceeded = limiter.too_many_attempts(key, window_sec, max_attempts)
        except Exception:
            exceeded = True
        if exceeded:
            raise HTTPException(status_code=429, detail="too many attempts")
        raise HTTPException(status_code=403, detail="invalid password")
    try:
        ttl = int(os.environ.get("ADMIN_TOKEN_TTL", "7200") or 7200)
    except Exception:
        ttl = 7200
    try:
        ip = getattr(request.client, 'host', None) or request.headers.get('x-forwarded-for', '').split(',')[0].strip() or 'unknown'
        get_admin_rate_limiter().reset(f"fastapi:{ip}")
    except Exception:
        pass
    tok = issue_admin_token(ttl_seconds=ttl)
    response = JSONResponse({"ok": True, "token": tok, "ttl": ttl})
    response.set_cookie(
        key=ADMIN_COOKIE_NAME,
        value=tok,
        httponly=True,
        samesite="lax",
        secure=bool(os.environ.get("COOKIE_SECURE", "").lower() in {"1", "true", "yes", "on"}),
        max_age=ttl,
    )
    return response


@router.get("/portal/{slug}/payroll/{year}/{month}", response_model=PayrollRowsResponse)
@router.get("/api/portal/{slug}/payroll/{year}/{month}", response_model=PayrollRowsResponse)
def api_get_payroll(
    slug: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    # Read access allows 'viewer' as well
    company = require_company_with_roles(slug, db, {"viewer", "payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    rec = (
        db.query(MonthlyPayroll)
        .filter(
            MonthlyPayroll.company_id == company.id,
            MonthlyPayroll.year == year,
            MonthlyPayroll.month == month,
        )
        .first()
    )
    if not rec:
        return {"ok": True, "rows": []}
    try:
        rows = json.loads(rec.rows_json or "[]")
    except Exception:
        rows = []
    return {"ok": True, "rows": rows}


@router.post("/portal/{slug}/calc/deductions", response_model=PayrollCalcResponse)
@router.post("/api/portal/{slug}/calc/deductions", response_model=PayrollCalcResponse)
def api_calc_deductions(
    slug: str,
    payload: PayrollCalcRequest,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    row = payload.row or {}
    try:
        year = int(payload.year)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid year")
    amounts, metadata = compute_deductions_service(db, company, row, year)
    return PayrollCalcResponse(amounts=amounts, metadata=metadata)


def _looks_like_int(s: str) -> bool:
    try:
        t = s.strip().replace(",", "")
        if t.startswith("-"):
            t = t[1:]
        return t.isdigit()
    except Exception:
        return False


def _parse_value(v: str):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return ""
    low = s.lower()
    if low in {"on", "true", "t", "yes", "y", "1"}:
        return True
    if _looks_like_int(s):
        try:
            return int(s.replace(",", ""))
        except Exception:
            pass
    # leave as string
    return s


def _parse_rows_from_form(form: dict) -> list[dict]:
    bucket: dict[int, dict] = {}
    for k, v in form.items():
        if not k.startswith("rows["):
            continue
        try:
            left, right = k.split("][", 1)
            idx = int(left[5:])
            field = right[:-1]
        except Exception:
            continue
        bucket.setdefault(idx, {})[field] = _parse_value(v)
    rows: list[dict] = []
    for idx in sorted(bucket.keys()):
        row = bucket[idx]
        # skip fully empty rows
        if not any(str(val or "").strip() for val in row.values()):
            continue
        rows.append(row)
    return rows


@router.post("/portal/{slug}/payroll/{year}/{month}", response_model=SimpleOkResponse)
async def api_save_payroll(
    slug: str,
    year: int,
    month: int,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, None, portal_cookie)
    # accept x-www-form-urlencoded or JSON
    rows: list[dict]
    ct = request.headers.get("content-type", "")
    if "application/json" in ct:
        payload = await request.json()
        rows = payload.get("rows") or []
        if not isinstance(rows, list):
            raise HTTPException(status_code=400, detail="invalid payload")
    else:
        form = await request.form()
        rows = _parse_rows_from_form(dict(form))

    rec = (
        db.query(MonthlyPayroll)
        .filter(
            MonthlyPayroll.company_id == company.id,
            MonthlyPayroll.year == year,
            MonthlyPayroll.month == month,
        )
        .first()
    )
    if rec and bool(getattr(rec, "is_closed", False)):
        raise HTTPException(status_code=400, detail="month is closed")

    # Prepare deterministic body hash for idempotency based on logical rows
    data = json.dumps(rows, ensure_ascii=False)
    body_hash = compute_body_hash({"rows": rows})

    def _produce():
        # Perform the actual save and return a JSON payload
        nonlocal rec
        if not rec:
            rec = MonthlyPayroll(company_id=company.id, year=year, month=month, rows_json=data)
            db.add(rec)
        else:
            rec.rows_json = data
        db.commit()
        try:
            audit_logger.info(
                "payroll_saved",
                extra={
                    "event": "payroll_saved",
                    "company_id": company.id,
                    "slug": company.slug,
                    "year": year,
                    "month": month,
                    "rows_count": len(rows) if isinstance(rows, list) else 0,
                },
            )
        except Exception:
            pass
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=company.id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(content=content, status_code=status)


# ------------------------------
# Fields config (parity with Flask JSON APIs)
# ------------------------------

def _load_include_map(db: Session, company: Company) -> dict[str, dict[str, bool]]:
    inc: dict[str, dict[str, bool]] = {"nhis": {}, "ei": {}}
    try:
        rows = db.query(FieldPref.field, FieldPref.ins_nhis, FieldPref.ins_ei).filter(FieldPref.company_id == company.id).all()
        for field, ins_nhis, ins_ei in rows:
            if bool(ins_nhis):
                inc["nhis"][field] = True
            if bool(ins_ei):
                inc["ei"][field] = True
    except Exception:
        # Backward compatible fallback
        pass
    return inc


@router.get("/portal/{slug}/fields/calc-config", response_model=FieldCalcConfigResponse)
@router.get("/api/portal/{slug}/fields/calc-config", response_model=FieldCalcConfigResponse)
def api_get_calc_config(
    slug: str,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    include = _load_include_map(db, company)
    return FieldCalcConfigResponse(include=FieldCalcInclude(**include))


@router.post("/portal/{slug}/fields/calc-config", response_model=SimpleOkResponse)
@router.post("/api/portal/{slug}/fields/calc-config", response_model=SimpleOkResponse)
def api_save_calc_config(
    slug: str,
    payload: FieldCalcConfigRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    inc = payload.include or {}
    nhis = inc.get("nhis") or {}
    ei = inc.get("ei") or {}
    if not isinstance(nhis, dict) or not isinstance(ei, dict):
        raise HTTPException(status_code=400, detail="invalid payload")
    nhis_keys = {k for k, v in nhis.items() if v}
    ei_keys = {k for k, v in ei.items() if v}
    body_hash = compute_body_hash({"type": "calc-config", "nhis": sorted(nhis_keys), "ei": sorted(ei_keys)})

    def _produce():
        # Upsert
        for key in nhis_keys | ei_keys:
            pref = db.query(FieldPref).filter(FieldPref.company_id == company.id, FieldPref.field == key).first()
            if not pref:
                pref = FieldPref(company_id=company.id, field=key)
                db.add(pref)
            pref.ins_nhis = key in nhis_keys
            pref.ins_ei = key in ei_keys
        # Reset others
        rows = db.query(FieldPref).filter(FieldPref.company_id == company.id).all()
        for p in rows:
            if p.field not in nhis_keys:
                p.ins_nhis = False
            if p.field not in ei_keys:
                p.ins_ei = False
        db.commit()
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=company.id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(status_code=status, content=content)


def _base_exemptions_from_env() -> dict:
    try:
        raw = os.environ.get("INS_BASE_EXEMPTIONS", "")
        if not raw:
            return {}
        return json.loads(raw) if isinstance(raw, str) else {}
    except Exception:
        return {}


@router.get("/portal/{slug}/fields/exempt-config", response_model=FieldExemptConfigResponse)
@router.get("/api/portal/{slug}/fields/exempt-config", response_model=FieldExemptConfigResponse)
def api_get_exempt_config(
    slug: str,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    ex: dict[str, FieldExemptEntry] = {}
    try:
        rows = db.query(FieldPref.field, FieldPref.exempt_enabled, FieldPref.exempt_limit).filter(FieldPref.company_id == company.id).all()
        for field, enabled, limit in rows:
            if bool(enabled) or int(limit or 0) > 0:
                ex[field] = FieldExemptEntry(enabled=bool(enabled), limit=int(limit or 0))
    except Exception:
        # Backward compatible fallback when schema differs
        pass
    return FieldExemptConfigResponse(exempt=ex, base=_base_exemptions_from_env())


@router.post("/portal/{slug}/fields/exempt-config", response_model=SimpleOkResponse)
@router.post("/api/portal/{slug}/fields/exempt-config", response_model=SimpleOkResponse)
def api_save_exempt_config(
    slug: str,
    payload: FieldExemptConfigRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    raw = payload.exempt or {}
    body_hash = compute_body_hash({
        "type": "exempt-config",
        "items": sorted([(k, bool(v.enabled), int(v.limit or 0)) for k, v in raw.items()]),
    })

    def _produce():
        for field, conf in raw.items():
            enabled = bool(conf.enabled)
            limit = int(conf.limit or 0)
            pref = db.query(FieldPref).filter(FieldPref.company_id == company.id, FieldPref.field == field).first()
            if not pref:
                pref = FieldPref(company_id=company.id, field=field)
                db.add(pref)
            pref.exempt_enabled = enabled
            pref.exempt_limit = limit
        db.commit()
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,
        company_id=company.id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(status_code=status, content=content)


@router.post("/portal/{slug}/fields/add", response_model=FieldAddResponse)
@router.post("/api/portal/{slug}/fields/add", response_model=FieldAddResponse)
def api_add_field(
    slug: str,
    payload: FieldAddRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    label = payload.label.strip()
    typ = payload.typ.strip()
    if not label:
        raise HTTPException(status_code=400, detail="label required")
    body_hash = compute_body_hash({"action": "add_field", "label": label, "typ": typ})

    def _produce():
        existing = db.query(ExtraField).filter(ExtraField.company_id == company.id, ExtraField.label == label).first()
        if existing:
            field_info = FieldInfo(name=existing.name, label=existing.label, typ=existing.typ)
            return {"ok": True, "field": field_info.dict(), "existed": True}, 200
        # Generate unique name from label
        base = label
        name = base
        i = 1
        while db.query(ExtraField).filter(ExtraField.company_id == company.id, ExtraField.name == name).first():
            i += 1
            name = f"{base}_{i}"
        ef = ExtraField(company_id=company.id, name=name, label=label, typ=typ)
        db.add(ef)
        db.commit()
        try:
            cleanup_duplicate_extra_fields(db, company)
        except Exception:
            pass
        return {"ok": True, "field": {"name": ef.name, "label": ef.label, "typ": ef.typ}, "existed": False}, 200

    content, status = maybe_idempotent_json(db, request, company_id=company.id, body_hash=body_hash, produce=_produce)
    return JSONResponse(status_code=status, content=content)


@router.post("/portal/{slug}/fields/delete", response_model=SimpleOkResponse)
@router.post("/api/portal/{slug}/fields/delete", response_model=SimpleOkResponse)
def api_delete_field(
    slug: str,
    payload: FieldDeleteRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    body_hash = compute_body_hash({"action": "delete_field", "name": name})

    def _produce():
        ef = db.query(ExtraField).filter(ExtraField.company_id == company.id, ExtraField.name == name).first()
        if not ef:
            return {"ok": False, "error": "not found"}, 404
        db.delete(ef)
        db.commit()
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(db, request, company_id=company.id, body_hash=body_hash, produce=_produce)
    return JSONResponse(status_code=status, content=content)


@router.post("/portal/{slug}/fields/group-config", response_model=FieldGroupConfigResponse)
@router.post("/api/portal/{slug}/fields/group-config", response_model=FieldGroupConfigResponse)
def api_save_group_config(
    slug: str,
    payload: FieldGroupConfigRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"payroll_manager", "company_admin", "admin"}, authorization, x_api_token, token, portal_cookie)
    group_map = payload.map or {}
    alias_map = payload.alias or {}
    body_hash = compute_body_hash({
        "type": "group-config",
        "map": sorted(group_map.items()),
        "alias": sorted(alias_map.items()),
    })

    def _produce():
        for field, grp in group_map.items():
            grp = (grp or "none").strip()
            pref = db.query(FieldPref).filter(FieldPref.company_id == company.id, FieldPref.field == field).first()
            if not pref:
                pref = FieldPref(company_id=company.id, field=field, group=grp)
                db.add(pref)
            else:
                pref.group = grp
        for field, alias in alias_map.items():
            alias = (alias or "").strip()
            pref = db.query(FieldPref).filter(FieldPref.company_id == company.id, FieldPref.field == field).first()
            if not pref:
                pref = FieldPref(company_id=company.id, field=field, alias=alias)
                db.add(pref)
            else:
                pref.alias = alias
        db.commit()
        try:
            cleanup_duplicate_extra_fields(db, company)
        except Exception:
            pass
        return FieldGroupConfigResponse().dict(), 200

    content, status = maybe_idempotent_json(db, request, company_id=company.id, body_hash=body_hash, produce=_produce)
    return JSONResponse(status_code=status, content=content)


# ------------------------------
# Prorate config (일할계산 적용 항목)
# ------------------------------

def _load_prorate_map(db: Session, company: Company) -> dict[str, bool]:
    try:
        rows = db.query(FieldPref).filter(FieldPref.company_id == company.id).all()
        out: dict[str, bool] = {}
        for p in rows:
            if bool(getattr(p, "prorate", False)):
                out[p.field] = True
        return out
    except Exception:
        # Backward-compat: if column does not exist yet, return empty map
        return {}


@router.get("/api/portal/{slug}/fields/prorate-config", response_model=FieldProrateConfigResponse)
@router.get("/portal/{slug}/fields/prorate-config", response_model=FieldProrateConfigResponse)
def api_get_prorate_config(
    slug: str,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    return FieldProrateConfigResponse(prorate=_load_prorate_map(db, company))


@router.post("/api/portal/{slug}/fields/prorate-config", response_model=SimpleOkResponse)
@router.post("/portal/{slug}/fields/prorate-config", response_model=SimpleOkResponse)
def api_save_prorate_config(
    slug: str,
    payload: FieldProrateConfigRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_admin_token: Optional[str] = Header(None),
    admin_cookie: Optional[str] = Cookie(None, alias=ADMIN_COOKIE_NAME),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    # Require company context for slug validation
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    # Admin-only mutation
    require_admin(authorization, x_admin_token, None, admin_cookie)
    m = payload.prorate or {}
    if not isinstance(m, dict):
        raise HTTPException(status_code=400, detail="invalid payload")
    keys = {k for k, v in m.items() if v}
    body_hash = compute_body_hash({"type": "prorate-config", "keys": sorted(keys)})

    def _produce():
        try:
            # Upsert selected
            for key in keys:
                pref = db.query(FieldPref).filter(FieldPref.company_id == company.id, FieldPref.field == key).first()
                if not pref:
                    pref = FieldPref(company_id=company.id, field=key)
                    db.add(pref)
                pref.prorate = True
            # Reset others
            rows = db.query(FieldPref).filter(FieldPref.company_id == company.id).all()
            for p in rows:
                if p.field not in keys:
                    p.prorate = False
            db.commit()
            return SimpleOkResponse().dict(), 200
        except Exception:
            return {"ok": False, "error": "prorate config unavailable; apply DB migration"}, 400

    content, status = maybe_idempotent_json(db, request, company_id=company.id, body_hash=body_hash, produce=_produce)
    return JSONResponse(status_code=status, content=content)


# ------------------------------
# Close / Open month
# ------------------------------

@router.post("/portal/{slug}/payroll/{year}/{month}/close", response_model=SimpleOkResponse)
def api_close_month(
    slug: str,
    year: int,
    month: int,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"company_admin", "admin"}, authorization, x_api_token, None, portal_cookie)
    rec = (
        db.query(MonthlyPayroll)
        .filter(
            MonthlyPayroll.company_id == company.id,
            MonthlyPayroll.year == year,
            MonthlyPayroll.month == month,
        )
        .first()
    )
    # Idempotency by key if provided
    body_hash = compute_body_hash({"action": "close", "slug": slug, "year": year, "month": month})

    def _produce():
        nonlocal rec
        if not rec:
            rec = MonthlyPayroll(company_id=company.id, year=year, month=month, rows_json="[]", is_closed=True)
            db.add(rec)
        else:
            rec.is_closed = True
        db.commit()
        try:
            audit_logger.info(
                "month_closed",
                extra={
                    "event": "month_closed",
                    "company_id": company.id,
                    "slug": company.slug,
                    "year": year,
                    "month": month,
                },
            )
        except Exception:
            pass
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,  # type: ignore[name-defined]
        company_id=company.id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(content=content, status_code=status)


@router.post("/portal/{slug}/payroll/{year}/{month}/open", response_model=SimpleOkResponse)
def api_open_month(
    slug: str,
    year: int,
    month: int,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company_with_roles(slug, db, {"company_admin", "admin"}, authorization, x_api_token, None, portal_cookie)
    rec = (
        db.query(MonthlyPayroll)
        .filter(
            MonthlyPayroll.company_id == company.id,
            MonthlyPayroll.year == year,
            MonthlyPayroll.month == month,
        )
        .first()
    )
    body_hash = compute_body_hash({"action": "open", "slug": slug, "year": year, "month": month})

    def _produce():
        if rec:
            rec.is_closed = False
            db.commit()
        try:
            audit_logger.info(
                "month_opened",
                extra={
                    "event": "month_opened",
                    "company_id": company.id,
                    "slug": company.slug,
                    "year": year,
                    "month": month,
                },
            )
        except Exception:
            pass
        return {"ok": True}, 200

    content, status = maybe_idempotent_json(
        db,
        request,  # type: ignore[name-defined]
        company_id=company.id,
        body_hash=body_hash,
        produce=_produce,
    )
    return JSONResponse(content=content, status_code=status)


# ------------------------------
# Export (basic workbook)
# ------------------------------
@router.get("/portal/{slug}/export/{year}/{month}")
def api_export(
    slug: str,
    year: int,
    month: int,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
    x_api_token: Optional[str] = Header(None),
    token: Optional[str] = None,
    portal_cookie: Optional[str] = Cookie(None, alias=PORTAL_COOKIE_NAME),
):
    company = require_company(slug, db, authorization, x_api_token, token, portal_cookie)
    # Optional signed link enforcement (does not break existing when secret not set)
    import os, hmac
    from hashlib import sha256
    secret = (os.environ.get("EXPORT_HMAC_SECRET") or "").strip()
    if secret:
        exp = (int)(request.query_params.get("exp") or 0)
        sig = (request.query_params.get("sig") or "").strip()
        if not exp or not sig:
            raise HTTPException(status_code=403, detail="missing signature")
        import time
        if exp < int(time.time()):
            raise HTTPException(status_code=403, detail="link expired")
        msg = f"{request.url.path}|{exp}|{company.id}".encode()
        ok = hmac.compare_digest(hmac.new(secret.encode(), msg, sha256).hexdigest(), sig)
        if not ok:
            raise HTTPException(status_code=403, detail="invalid signature")
    rec = (
        db.query(MonthlyPayroll)
        .filter(
            MonthlyPayroll.company_id == company.id,
            MonthlyPayroll.year == year,
            MonthlyPayroll.month == month,
        )
        .first()
    )
    if not rec:
        raise HTTPException(status_code=400, detail="no data to export")
    try:
        rows = json.loads(rec.rows_json or "[]")
    except Exception:
        rows = []
    # Build workbook identical to Flask export using shared exporter
    from core.exporter import build_salesmap_workbook_stream_spooled as build_salesmap_workbook_stream
    from core.schema import DEFAULT_COLUMNS
    # Build all_columns = DEFAULT + extras
    extras = db.query(ExtraField).filter(ExtraField.company_id == company.id).order_by(ExtraField.position.asc(), ExtraField.id.asc()).all()
    all_columns = list(DEFAULT_COLUMNS) + [(e.name, e.label, e.typ or 'number') for e in extras]
    # group/alias prefs
    rows_pref = db.query(FieldPref).filter(FieldPref.company_id == company.id).all()
    gp = {}
    ap = {}
    for p in rows_pref:
        if p.group and p.group != "none":
            gp[p.field] = p.group
        if p.alias:
            ap[p.field] = p.alias
    bio = build_salesmap_workbook_stream(
        company_slug=company.slug,
        year=year,
        month=month,
        rows=rows,
        all_columns=all_columns,
        group_prefs=gp,
        alias_prefs=ap,
    )
    filename = f"{company.slug}_{year}-{month:02d}_세일즈맵.xlsx"
    encoded = quote(filename)
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    # Audit (DB) best-effort
    try:
        ip = getattr(request.client, 'host', None) or request.headers.get('x-forwarded-for', '').split(',')[0].strip() or ''
        ua = request.headers.get('user-agent', '')
        record_event(
            db,
            actor=f"company:{company.slug}",
            action='export_download',
            resource=str(request.url.path),
            company_id=company.id,
            ip=ip,
            ua=ua,
            meta={"year": year, "month": month},
        )
    except Exception:
        pass
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
def require_company(
    slug: str,
    db: Session,
    authorization: Optional[str] = None,
    x_api_token: Optional[str] = None,
    token: Optional[str] = None,
    portal_cookie: Optional[str] = None,
) -> Company:
    tok = extract_token(authorization, x_api_token, token, portal_cookie)
    if not tok:
        raise HTTPException(status_code=403, detail="missing token")
    company = authenticate_company(db, slug, tok)
    if not company:
        if not get_company_by_slug(db, slug):
            raise HTTPException(status_code=404, detail="company not found")
        raise HTTPException(status_code=403, detail="invalid token")
    return company


def require_company_with_roles(
    slug: str,
    db: Session,
    required_roles: set[str],
    authorization: Optional[str] = None,
    x_api_token: Optional[str] = None,
    token: Optional[str] = None,
    portal_cookie: Optional[str] = None,
) -> Company:
    tok = extract_token(authorization, x_api_token, token, portal_cookie)
    if not tok:
        raise HTTPException(status_code=403, detail="missing token")
    company = authenticate_company(db, slug, tok)
    if not company:
        if not get_company_by_slug(db, slug):
            raise HTTPException(status_code=404, detail="company not found")
        raise HTTPException(status_code=403, detail="invalid token")
    # roles check
    try:
        roles = set(token_roles(tok, is_admin=False))
        if roles and required_roles and roles.isdisjoint(required_roles):
            raise HTTPException(status_code=403, detail="forbidden")
    except HTTPException:
        raise
    except Exception:
        pass
    return company


@router.get("/meta")
def meta():
    settings = get_settings()
    return {
        "app_version": settings.app_version,
        "git_sha": settings.git_sha or "",
        "build_ts": settings.build_ts or "",
    }


def create_app() -> FastAPI:
    # Optional observability (JSON logs + Sentry) enabled by env vars
    maybe_enable_json_logging()
    init_sentry()
    application = FastAPI(title="Payroll API (FastAPI)", lifespan=lifespan)

    origins_env = (os.environ.get("API_CORS_ORIGINS") or "").strip()
    if origins_env:
        origins = [o.strip() for o in origins_env.split(",") if o.strip()]
    else:
        api_base = (os.environ.get("API_BASE_URL") or "").strip()
        if api_base:
            origins = [api_base]
        else:
            origins = [
                "http://localhost:5000",
                "http://127.0.0.1:5000",
                "http://localhost:8000",
                "http://127.0.0.1:8000",
            ]

    application.add_middleware(
        CORSMiddleware,
        allow_origins=origins,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @application.middleware("http")
    async def _request_id_mw(request, call_next):
        rid = request.headers.get("X-Request-ID") or uuid.uuid4().hex
        set_request_id(rid)
        resp = await call_next(request)
        try:
            resp.headers["X-Request-ID"] = rid
        except Exception:
            pass
        return resp

    application.include_router(router)
    register_exception_handlers(application)

    def custom_openapi():
        if application.openapi_schema:
            return application.openapi_schema
        openapi_schema = get_openapi(
            title=application.title,
            version="1.0.0",
            description="Payroll API with problem+json and idempotency support.",
            routes=application.routes,
        )
        comps = openapi_schema.setdefault("components", {})
        security = comps.setdefault("securitySchemes", {})
        security.setdefault(
            "AdminToken",
            {"type": "apiKey", "in": "header", "name": "X-Admin-Token", "description": "Admin token or Authorization Bearer"},
        )
        security.setdefault(
            "CompanyToken",
            {"type": "apiKey", "in": "header", "name": "X-API-Token", "description": "Company portal token or Authorization Bearer"},
        )
        params = comps.setdefault("parameters", {})
        params.setdefault(
            "IdempotencyKey",
            {
                "name": "Idempotency-Key",
                "in": "header",
                "required": False,
                "schema": {"type": "string"},
                "description": "Provide to make mutation requests idempotent.",
            },
        )
        schemas = comps.setdefault("schemas", {})
        schemas.setdefault(
            "ProblemDetails",
            {
                "type": "object",
                "properties": {
                    "type": {"type": "string"},
                    "title": {"type": "string"},
                    "status": {"type": "integer"},
                    "detail": {"type": "string"},
                    "instance": {"type": "string"},
                    "request_id": {"type": "string"},
                },
                "example": {
                    "type": "about:blank",
                    "title": "Conflict",
                    "status": 409,
                    "detail": "idempotency key conflict",
                    "instance": "/api/v1/admin/company/1/reset-code",
                    "request_id": "c7f2d9e2a1b34f7f8a90d1",
                },
            },
        )
        # Enrich operations: add Idempotency-Key parameter and problem+json error responses to mutating verbs
        try:
            paths = openapi_schema.get("paths", {})
            for path, ops in list(paths.items()):
                if not isinstance(ops, dict):
                    continue
                for method, op in list(ops.items()):
                    if method.lower() not in {"post", "put", "patch", "delete"}:
                        continue
                    # Attach Idempotency-Key header parameter (skip if exists)
                    params_list = op.setdefault("parameters", [])
                    has_idem = any(p.get("$ref") == "#/components/parameters/IdempotencyKey" or p.get("name") == "Idempotency-Key" for p in params_list if isinstance(p, dict))
                    if not has_idem:
                        params_list.append({"$ref": "#/components/parameters/IdempotencyKey"})
                    # Standard problem+json responses
                    responses = op.setdefault("responses", {})
                    for code in ("400", "401", "403", "404", "409"):
                        if code not in responses:
                            responses[code] = {
                                "description": "Error",
                                "content": {
                                    "application/problem+json": {
                                        "schema": {"$ref": "#/components/schemas/ProblemDetails"}
                                    }
                                },
                            }
        except Exception:
            pass
        # Enrich with examples and security hints
        try:
            paths = openapi_schema.get("paths", {})
            # Example: Admin create company (form)
            if "/admin/company/new" in paths and "post" in paths["/admin/company/new"]:
                op = paths["/admin/company/new"]["post"]
                rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/x-www-form-urlencoded", {})
                rb.setdefault(
                    "example",
                    {
                        "name": "테스트회사",
                        "slug": "acme",
                    },
                )
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True, "company": {"id": 1, "name": "테스트회사", "slug": "acme", "created_at": "2025-01-01T00:00:00Z"}, "access_code": "abcd1234"},
                ))

            # Example: Save payroll (JSON)
            for p in ("/portal/{slug}/payroll/{year}/{month}", "/api/portal/{slug}/payroll/{year}/{month}"):
                if p in paths and "post" in paths[p]:
                    op = paths[p]["post"]
                    rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                    rb.setdefault(
                        "example",
                        {
                            "rows": [
                                {"사원코드": "E01", "사원명": "홍길동", "기본급": 2000000, "식대": 200000, "소득세": 100000, "지방소득세": 10000}
                            ]
                        },
                    )
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True},
                    ))

            # Example: Withholding cells page (query)
            if "/admin/tax/withholding/cells" in paths and "get" in paths["/admin/tax/withholding/cells"]:
                op = paths["/admin/tax/withholding/cells"]["get"]
                # Attach examples to parameters
                for prm in op.get("parameters", []):
                    if prm.get("name") == "year":
                        prm.setdefault("example", 2025)
                    if prm.get("name") == "dep":
                        prm.setdefault("example", 1)
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True, "items": [{"dependents": 1, "wage": 2000000, "tax": 110000}], "has_more": False},
                ))

            # Examples for field config endpoints
            for p in ("/api/portal/{slug}/fields/calc-config", "/portal/{slug}/fields/calc-config"):
                if p in paths and "post" in paths[p]:
                    op = paths[p]["post"]
                    rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                    rb.setdefault("example", {"include": {"nhis": {"기본급": True}, "ei": {"기본급": True}}})
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True},
                    ))
            for p in ("/api/portal/{slug}/fields/exempt-config", "/portal/{slug}/fields/exempt-config"):
                if p in paths and "post" in paths[p]:
                    op = paths[p]["post"]
                    rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                    rb.setdefault("example", {"exempt": {"식대": {"enabled": True, "limit": 100000}}})
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True, "exempt": {"식대": {"enabled": True, "limit": 100000}}},
                    ))
            for p in ("/api/portal/{slug}/fields/group-config", "/portal/{slug}/fields/group-config"):
                if p in paths and "post" in paths[p]:
                    op = paths[p]["post"]
                    rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                    rb.setdefault("example", {"map": {"기본급": "earn", "소득세": "deduct"}, "alias": {"기본급": "Base Salary"}})
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True},
                    ))
            for p in ("/api/portal/{slug}/fields/prorate-config", "/portal/{slug}/fields/prorate-config"):
                if p in paths and "post" in paths[p]:
                    op = paths[p]["post"]
                    rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                    rb.setdefault("example", {"prorate": {"기본급": True, "상여": False}})
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True},
                    ))

            # Examples for admin company reset-code / rotate-token-key (POST)
            p = "/admin/company/{company_id}/reset-code"
            if p in paths and "post" in paths[p]:
                op = paths[p]["post"]
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True, "company_id": 1, "access_code": "1a2b3c4d"},
                ))
                # Error examples
                err = op.setdefault("responses", {}).setdefault("403", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err.setdefault("example", {"type": "about:blank", "title": "Forbidden", "status": 403, "detail": "forbidden", "instance": "/api/v1/admin/company/1/reset-code"})
                err404 = op.setdefault("responses", {}).setdefault("404", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err404.setdefault("example", {"type": "about:blank", "title": "Not Found", "status": 404, "detail": "not found", "instance": "/api/v1/admin/company/999/reset-code"})
                err409 = op.setdefault("responses", {}).setdefault("409", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err409.setdefault("example", {"type": "about:blank", "title": "Conflict", "status": 409, "detail": "idempotency key conflict", "instance": "/api/v1/admin/company/1/reset-code"})
            p = "/admin/company/{company_id}/rotate-token-key"
            if p in paths and "post" in paths[p]:
                op = paths[p]["post"]
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True},
                ))
                err = op.setdefault("responses", {}).setdefault("403", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err.setdefault("example", {"type": "about:blank", "title": "Forbidden", "status": 403, "detail": "forbidden", "instance": "/api/v1/admin/company/1/rotate-token-key"})
                err404 = op.setdefault("responses", {}).setdefault("404", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err404.setdefault("example", {"type": "about:blank", "title": "Not Found", "status": 404, "detail": "not found", "instance": "/api/v1/admin/company/999/rotate-token-key"})
                err409 = op.setdefault("responses", {}).setdefault("409", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err409.setdefault("example", {"type": "about:blank", "title": "Conflict", "status": 409, "detail": "idempotency key conflict", "instance": "/api/v1/admin/company/1/rotate-token-key"})

            # Example for admin impersonate-token (GET)
            p = "/admin/company/{company_id}/impersonate-token"
            if p in paths and "get" in paths[p]:
                op = paths[p]["get"]
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True, "slug": "acme", "token": "eyJhbGciOi..."},
                ))
                err = op.setdefault("responses", {}).setdefault("403", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err.setdefault("example", {"type": "about:blank", "title": "Forbidden", "status": 403, "detail": "forbidden", "instance": "/api/v1/admin/company/1/impersonate-token"})
                err404 = op.setdefault("responses", {}).setdefault("404", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err404.setdefault("example", {"type": "about:blank", "title": "Not Found", "status": 404, "detail": "not found", "instance": "/api/v1/admin/company/999/impersonate-token"})

            # Examples: paginated listings
            p = "/admin/companies/page"
            if p in paths and "get" in paths[p]:
                op = paths[p]["get"]
                for prm in op.get("parameters", []):
                    if prm.get("name") == "limit":
                        prm.setdefault("example", 20)
                    if prm.get("name") == "order":
                        prm.setdefault("example", "desc")
                    if prm.get("name") == "cursor":
                        prm.setdefault("example", "eyJpZCI6MiwgImNyZWF0ZWRfYXQiOiAiMjAyNS0wMS0wMlQwMDowMDowMFoiLCJvcmRlciI6ICJkZXNjIn0=")
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {
                        "ok": True,
                        "items": [
                            {"id": 2, "name": "베타", "slug": "beta", "created_at": "2025-01-02T00:00:00Z"},
                            {"id": 1, "name": "알파", "slug": "alpha", "created_at": "2025-01-01T00:00:00Z"},
                        ],
                        "has_more": False,
                        "next_cursor": None,
                    },
                ))
                err400 = op.setdefault("responses", {}).setdefault("400", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err400.setdefault("example", {"type": "about:blank", "title": "Bad Request", "status": 400, "detail": "invalid cursor", "instance": "/api/v1/admin/companies/page"})
                err403 = op.setdefault("responses", {}).setdefault("403", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err403.setdefault("example", {"type": "about:blank", "title": "Forbidden", "status": 403, "detail": "forbidden", "instance": "/api/v1/admin/companies/page"})

            p = "/admin/company/{company_id}/payrolls/page"
            if p in paths and "get" in paths[p]:
                op = paths[p]["get"]
                for prm in op.get("parameters", []):
                    if prm.get("name") == "year":
                        prm.setdefault("example", 2025)
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {
                        "ok": True,
                        "items": [
                            {"id": 10, "year": 2025, "month": 10, "is_closed": False, "updated_at": "2025-10-01T12:00:00Z"},
                            {"id": 9, "year": 2025, "month": 9, "is_closed": True, "updated_at": "2025-09-30T18:34:00Z"},
                        ],
                        "has_more": False,
                        "next_cursor": None,
                    },
                ))
                err400 = op.setdefault("responses", {}).setdefault("400", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err400.setdefault("example", {"type": "about:blank", "title": "Bad Request", "status": 400, "detail": "invalid cursor", "instance": "/api/v1/admin/company/1/payrolls/page"})
                err403 = op.setdefault("responses", {}).setdefault("403", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err403.setdefault("example", {"type": "about:blank", "title": "Forbidden", "status": 403, "detail": "forbidden", "instance": "/api/v1/admin/company/1/payrolls/page"})
                err404 = op.setdefault("responses", {}).setdefault("404", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err404.setdefault("example", {"type": "about:blank", "title": "Not Found", "status": 404, "detail": "not found", "instance": "/api/v1/admin/company/999/payrolls/page"})

            p = "/admin/company/{company_id}/extra-fields/page"
            if p in paths and "get" in paths[p]:
                op = paths[p]["get"]
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {
                        "ok": True,
                        "items": [
                            {"id": 1, "name": "식대", "label": "식대", "typ": "number", "position": 10},
                            {"id": 2, "name": "상여", "label": "상여", "typ": "number", "position": 20},
                        ],
                        "has_more": False,
                        "next_cursor": None,
                    },
                ))
                err403 = op.setdefault("responses", {}).setdefault("403", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err403.setdefault("example", {"type": "about:blank", "title": "Forbidden", "status": 403, "detail": "forbidden", "instance": "/api/v1/admin/company/1/extra-fields/page"})
                err404 = op.setdefault("responses", {}).setdefault("404", {}).setdefault("content", {}).setdefault("application/problem+json", {})
                err404.setdefault("example", {"type": "about:blank", "title": "Not Found", "status": 404, "detail": "not found", "instance": "/api/v1/admin/company/999/extra-fields/page"})

            # Policy endpoints
            p = "/admin/policy"
            if p in paths and "get" in paths[p]:
                op = paths[p]["get"]
                for prm in op.get("parameters", []):
                    if prm.get("name") == "year":
                        prm.setdefault("example", 2025)
                    if prm.get("name") == "company_id":
                        prm.setdefault("example", 1)
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True, "policy": {"local_tax": {"rate": 0.1, "round_to": 10, "rounding": "round"}}},
                ))
            if p in paths and "post" in paths[p]:
                op = paths[p]["post"]
                rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                rb.setdefault("example", {"local_tax": {"rate": 0.1, "round_to": 10, "rounding": "round"}})
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {"ok": True},
                ))
            # UI prefs endpoints examples
            for p in ("/api/portal/{slug}/ui-prefs", "/portal/{slug}/ui-prefs"):
                if p in paths and "get" in paths[p]:
                    op = paths[p]["get"]
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True, "values": {"table.columnWidths": {"기본급": 180, "사원명": 140}}},
                    ))
                if p in paths and "post" in paths[p]:
                    op = paths[p]["post"]
                    rb = op.setdefault("requestBody", {}).setdefault("content", {}).setdefault("application/json", {})
                    rb.setdefault("example", {"values": {"table.columnWidths": {"기본급": 180}}})
                    (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                        "example",
                        {"ok": True},
                    ))
            # Policy history endpoint example
            p = "/admin/policy/history"
            if p in paths and "get" in paths[p]:
                op = paths[p]["get"]
                for prm in op.get("parameters", []):
                    if prm.get("name") == "year":
                        prm.setdefault("example", 2025)
                    if prm.get("name") == "company_id":
                        prm.setdefault("example", 1)
                (op.setdefault("responses", {}).setdefault("200", {}).setdefault("content", {}).setdefault("application/json", {}).setdefault(
                    "example",
                    {
                        "ok": True,
                        "items": [
                            {"id": 1, "ts": "2025-10-29T12:00:00Z", "actor": "admin", "company_id": 1, "year": 2025, "old": {}, "new": {"local_tax": {"round_to": 10}}}
                        ],
                        "has_more": False,
                        "next_cursor": None,
                    },
                ))
        except Exception:
            pass

        application.openapi_schema = openapi_schema
        return application.openapi_schema

    application.openapi = custom_openapi
    return application


app = create_app()
# Simple helper to expose withholding tax computation for tests and internal callers
def compute_withholding_tax(
    db: Session, *, year: int, dependents: int, wage: int
) -> int:
    return int(compute_withholding_tax_service(db, year, dependents, wage) or 0)
# Lightweight admin guard usable both in FastAPI routes and direct calls
def require_admin(
    authorization: Optional[str] = None,
    x_admin_token: Optional[str] = None,
    query_token: Optional[str] = None,
    admin_cookie: Optional[str] = None,
) -> None:
    tok = extract_token(authorization, x_admin_token, query_token, admin_cookie)
    if not tok or not authenticate_admin(tok):
        raise HTTPException(status_code=403, detail="forbidden")
    # roles enforcement: must contain 'admin' if roles present
    try:
        roles = set(token_roles(tok, is_admin=True))
        if roles and ("admin" not in roles):
            raise HTTPException(status_code=403, detail="forbidden")
    except HTTPException:
        raise
    except Exception:
        pass
