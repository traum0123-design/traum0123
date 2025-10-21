from __future__ import annotations

import datetime as dt
from typing import Optional

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from payroll_shared.models import FieldPref, WithholdingCell
from payroll_shared.settings import get_settings

from ..app.extensions import db_session
from ..dao import companies as companies_dao
from ..services import companies as company_service
from ..services.extra_fields import ensure_defaults
from ..services.rate_limit import admin_login_key, limiter


bp = Blueprint("admin", __name__, template_folder="../templates")


def _get_session() -> Session:
    return db_session()


def _admin_required():
    if not session.get("is_admin"):
        return redirect(url_for("admin.login"))
    return None


@bp.get("/login")
def login():
    return render_template("admin_login.html")


@bp.post("/login")
def login_post():
    password = request.form.get("password", "")
    rate_limiter = limiter()
    key = admin_login_key(request)
    max_attempts = int(getattr(get_settings(), "admin_login_max_attempts", 10) or 10)
    window_seconds = int(getattr(get_settings(), "admin_login_window", 600) or 600)
    if not company_service.verify_admin_password(password):
        exceeded = False
        try:
            exceeded = rate_limiter.too_many_attempts(key, window_seconds, max_attempts)
        except Exception:
            exceeded = True
        if exceeded:
            flash("시도 횟수가 많습니다. 잠시 후 다시 시도하세요.", "error")
        else:
            flash("관리자 비밀번호가 올바르지 않습니다.", "error")
        return redirect(url_for("admin.login"))

    try:
        rate_limiter.reset(key)
    except Exception:
        pass
    session["is_admin"] = True
    return redirect(url_for("admin.index"))


@bp.get("/logout")
def logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin.login"))


@bp.get("/")
def index():
    guard = _admin_required()
    if guard:
        return guard
    s = _get_session()
    companies = companies_dao.list_companies(s)
    try:
        rows = s.execute(
            select(
                WithholdingCell.year,
                func.count(),
                func.min(WithholdingCell.wage),
                func.max(WithholdingCell.wage),
            )
            .group_by(WithholdingCell.year)
            .order_by(WithholdingCell.year.desc())
        ).all()
        wh_counts = [
            {
                "year": row[0],
                "count": row[1],
                "min_wage": row[2],
                "max_wage": row[3],
            }
            for row in rows
        ]
    except Exception:
        wh_counts = []
    return render_template("admin_index.html", companies=companies, wh_counts=wh_counts)


@bp.post("/company/new")
def company_new():
    guard = _admin_required()
    if guard:
        return guard
    name = request.form.get("name", "").strip()
    slug = request.form.get("slug", "").strip().lower()
    if not name or not slug:
        flash("회사명과 슬러그를 모두 입력하세요.", "error")
        return redirect(url_for("admin.index"))
    s = _get_session()
    if companies_dao.get_by_slug(s, slug):
        flash("이미 사용중인 슬러그입니다.", "error")
        return redirect(url_for("admin.index"))
    company, code = company_service.create_company(s, name, slug)
    ensure_defaults(s, company)
    session.setdefault("new_codes", {})
    session["new_codes"][str(company.id)] = code
    session.modified = True
    return redirect(url_for("admin.company_detail", company_id=company.id))


@bp.get("/company/<int:company_id>")
def company_detail(company_id: int):
    guard = _admin_required()
    if guard:
        return guard
    s = _get_session()
    company = companies_dao.get_by_id(s, company_id)
    if not company:
        flash("회사를 찾을 수 없습니다.", "error")
        return redirect(url_for("admin.index"))
    new_codes = session.get("new_codes", {})
    new_code = new_codes.pop(str(company_id), None)
    session["new_codes"] = new_codes
    session.modified = True
    portal_login_url = url_for("portal.login", slug=company.slug, _external=False)
    return render_template(
        "admin_company_detail.html",
        company=company,
        new_code=new_code,
        portal_login_url=portal_login_url,
    )


@bp.post("/company/<int:company_id>/reset-code")
def company_reset_code(company_id: int):
    guard = _admin_required()
    if guard:
        return guard
    s = _get_session()
    company = companies_dao.get_by_id(s, company_id)
    if not company:
        flash("회사를 찾을 수 없습니다.", "error")
        return redirect(url_for("admin.index"))
    code = company_service.rotate_company_access(s, company)
    session.setdefault("new_codes", {})
    session["new_codes"][str(company_id)] = code
    session.modified = True
    return redirect(url_for("admin.company_detail", company_id=company_id))


@bp.get("/company/<int:company_id>/impersonate")
def company_impersonate(company_id: int):
    guard = _admin_required()
    if guard:
        return guard
    year = request.args.get("year")
    month = request.args.get("month")
    s = _get_session()
    company = companies_dao.get_by_id(s, company_id)
    if not company:
        flash("회사를 찾을 수 없습니다.", "error")
        return redirect(url_for("admin.index"))
    session["company_id"] = company_id
    session["company_slug"] = company.slug
    if year and month:
        try:
            year = int(year); month = int(month)
            return redirect(url_for("portal.edit_payroll", slug=company.slug, year=year, month=month))
        except Exception:
            pass
    return redirect(url_for("portal.home", slug=company.slug))


@bp.get("/tax/withholding")
def withholding_index():
    guard = _admin_required()
    if guard:
        return guard
    s = _get_session()
    try:
        years = s.execute(
            select(
                WithholdingCell.year,
                func.count(),
                func.min(WithholdingCell.wage),
                func.max(WithholdingCell.wage),
            )
            .group_by(WithholdingCell.year)
            .order_by(WithholdingCell.year.desc())
        ).all()
    except Exception:
        years = []
    return render_template("admin_withholding.html", years=years)


@bp.post("/tax/withholding/import")
def withholding_import():
    guard = _admin_required()
    if guard:
        return guard
    file = request.files.get("file")
    year = request.form.get("year", type=int)
    if not file or not year:
        flash("연도와 파일을 올바르게 입력하세요.", "error")
        return redirect(url_for("admin.withholding_index"))
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file, data_only=True)
        ws = wb.active
        header_row: Optional[int] = None
        dep_cols = {}
        for r in range(1, 15):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            tmp = {}
            for c, v in enumerate(row_vals[1:], start=2):
                try:
                    tmp[c] = int(str(v).strip().replace(",", ""))
                except Exception:
                    continue
            if len(tmp) >= 2:
                header_row = r
                dep_cols = tmp
                break
        if header_row is None:
            raise ValueError("의존가족수 헤더를 찾을 수 없습니다.")
        data = []
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            try:
                wage = int(float(str(v).replace(",", "").strip()))
            except Exception:
                if data:
                    break
                continue
            for c, dep in dep_cols.items():
                tv = ws.cell(row=r, column=c).value
                try:
                    tax = int(float(str(tv).replace(",", "").strip())) if tv not in (None, "") else 0
                except Exception:
                    tax = 0
                data.append((year, dep, wage, tax))
        s = _get_session()
        s.query(WithholdingCell).filter(WithholdingCell.year == year).delete()
        for y, dep, wage, tax in data:
            s.add(WithholdingCell(year=y, dependents=dep, wage=wage, tax=tax))
        s.commit()
        flash(f"간이세액표 {year}년 {len(data)}건을 저장했습니다.", "success")
    except Exception as exc:
        flash(f"가져오기 실패: {exc}", "error")
    return redirect(url_for("admin.withholding_index"))


@bp.get("/tax/withholding/sample")
def withholding_sample():
    guard = _admin_required()
    if guard:
        return guard
    year = request.args.get("year", type=int)
    dependents = request.args.get("dep", type=int)
    wage = request.args.get("wage", type=int)
    if not (year and dependents is not None and wage is not None):
        return {"ok": False, "error": "year/dep/wage 필요"}, 400
    s = _get_session()
    cell = companies_withholding_lookup(s, year, dependents, wage)
    if not cell:
        return {"ok": True, "year": year, "dep": dependents, "wage": wage, "tax": 0, "local_tax": 0}
    return {
        "ok": True,
        "year": year,
        "dep": dependents,
        "wage": wage,
        "tax": int(cell.tax),
        "local_tax": int(round(int(cell.tax) * 0.1)),
    }


def companies_withholding_lookup(session: Session, year: int, dependents: int, wage: int):
    return (
        session.query(WithholdingCell)
        .filter(
            WithholdingCell.year == year,
            WithholdingCell.dependents == dependents,
            WithholdingCell.wage <= wage,
        )
        .order_by(WithholdingCell.wage.desc())
        .first()
    )
